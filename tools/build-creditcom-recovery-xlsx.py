#!/usr/bin/env python3
"""Build the Credit.com Recovery Opportunity workbook.

Every number is sourced from the corrected rederive-v2 pipeline (2026-06-10):
  - live credit_portfolio:stats (mdl-business.vercel.app/api/credit-portfolio)
  - cr_db.db full-population SQL scans (settlement_cohort_sizes.json, business-directions audit)
  - casepeople sharded index (production KV)
  - data/settlements/open-settlements-2026-06.json + recovery-sizing-2026-06-10.md
  - PACER indexes in data/pacer-cases/

Run:  /usr/bin/python3 tools/build-creditcom-recovery-xlsx.py
Out:  data/credit-com-report/Creditcom-Recovery-Opportunity-2026-06-10.xlsx
"""
import json
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("/Users/stef/MDL Business/data/credit-com-report/Creditcom-Recovery-Opportunity-2026-06-10.xlsx")
OUT.parent.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------- styles
NAVY = "1F3864"; BLUE = "2E74B5"; LTBLUE = "DEEAF6"; GREEN = "C6E0B4"
YELLOW = "FFE699"; ORANGE = "F8CBAD"; GREY = "F2F2F2"; RED = "FFC7CE"
H1 = Font(name="Calibri", size=16, bold=True, color=NAVY)
H2 = Font(name="Calibri", size=12, bold=True, color=NAVY)
HDR = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
B = Font(name="Calibri", size=10, bold=True)
N = Font(name="Calibri", size=10)
SM = Font(name="Calibri", size=9, italic=True, color="595959")
HDR_FILL = PatternFill("solid", fgColor=NAVY)
SUB_FILL = PatternFill("solid", fgColor=LTBLUE)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAPC = Alignment(wrap_text=True, vertical="top", horizontal="center")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)
MONEY = '"$"#,##0'
MONEYM = '"$"#,##0.0,,"M"'
INT = "#,##0"

def sheet(wb, title, widths):
    ws = wb.create_sheet(title)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return ws

def header_row(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HDR; cell.fill = HDR_FILL; cell.alignment = WRAPC; cell.border = THIN
    return row + 1

def put(ws, r, c, v, font=N, fill=None, fmt=None, align=WRAP):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = font; cell.alignment = align; cell.border = THIN
    if fill: cell.fill = fill
    if fmt: cell.number_format = fmt
    return cell

def title_block(ws, text, sub, ncols):
    ws.cell(row=1, column=1, value=text).font = H1
    ws.cell(row=2, column=1, value=sub).font = SM
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)

# ---------------------------------------------------------------- PACER mining
# Open putative-class candidates (NOS 480, open or ran >1.5yr) and open national
# TCPA dockets (NOS 485), matched by title to the defendants our cohorts cover.
PACER = Path("/Users/stef/MDL Business/data/pacer-cases")
COHORTS = {  # production casepeople sharded index (rederive v2)
    "capital one": 1712966, "santander": 556939, "synchrony": 549694,
    "navient": 451340, "credit one": 403156, "wells fargo": 392187,
    "bank of america": 333611, "midland": 241303, "exeter": 235993,
    "westlake": 200689,
}
FDCPA_DEFS = [
    ("LVNV Funding", "lvnv"), ("Midland Credit / Midland Funding", "midland"),
    ("Portfolio Recovery Associates", "portfolio recovery"),
    ("Santander Consumer USA", "santander"), ("Credit One Bank", "credit one"),
    ("Westlake Financial / Westlake Portfolio", "westlake"),
    ("Jefferson Capital Systems", "jefferson capital"), ("Ally Financial", "ally"),
    ("Transworld Systems", "transworld"), ("I.C. System", "i.c. system"),
    ("Credit Acceptance", "credit acceptance"), ("Exeter Finance", "exeter"),
    ("First Premier Bank", "first premier"), ("OneMain Financial", "onemain"),
    ("Capital One", "capital one"), ("National Credit Adjusters", "national credit adjusters"),
]
TCPA_DEFS = [
    ("Capital One", "capital one"), ("Credit One Bank", "credit one"),
    ("Westlake Services / Portfolio Mgmt", "westlake"),
    ("Portfolio Recovery Associates", "portfolio recovery"),
    ("Wells Fargo", "wells fargo"), ("Synchrony Bank", "synchrony"),
    ("Santander Consumer USA", "santander"), ("Bank of America", "bank of america"),
    ("Citibank", "citibank"), ("Ally Financial", "ally"),
]

def mine(path, defs, court_key, docket_key):
    try:
        cases = json.load(open(path))
    except Exception:
        return {}, 0
    open_cases = [c for c in cases if c.get("status") == "open"]
    out = defaultdict(list)
    for c in open_cases:
        t = c.get("caseTitle", "").lower()
        for label, pat in defs:
            if pat in t:
                out[label].append(c); break
    for label in out:
        out[label].sort(key=lambda c: c.get("dateFiled", ""), reverse=True)
    return out, len(open_cases)

fdcpa_open, fdcpa_total_open = mine(PACER / "_candidates.json", FDCPA_DEFS, "courtId", "caseNumberFull")
tcpa_open, tcpa_total_open = mine(PACER / "_tcpa_index.json", TCPA_DEFS, "court", "docket_number")

def fmt_examples(cases, n=2, court_key="courtId", docket_key="caseNumberFull"):
    out = []
    for c in cases[:n]:
        out.append(f"{c['caseTitle']} ({c.get(court_key) or c.get('court')} {c.get(docket_key) or c.get('docket_number')}, filed {c.get('dateFiled')})")
    return "; ".join(out)

try:
    breach_mdls = json.load(open(PACER / "_breach_mdl_index.json"))
except Exception:
    breach_mdls = []
mdl_agg = defaultdict(lambda: [0, 0])
for c in breach_mdls:
    mdl_agg[(c["mdl"], c["mdlName"])][0] += 1
    if c.get("status") == "open":
        mdl_agg[(c["mdl"], c["mdlName"])][1] += 1

wb = openpyxl.Workbook()
wb.remove(wb.active)

# ================================================================ 1. EXEC SUMMARY
ws = sheet(wb, "Executive Summary", [38, 16, 16, 16, 70])
title_block(ws, "Credit.com Data — Legal Recovery Opportunity",
            "Prepared 2026-06-10 · All figures from the corrected (v2) full-population derivation of the LEX + CCOM dataset · Companion sheets carry per-item sourcing", 5)

r = 4
put(ws, r, 1, "THE ASSET", H2); r += 1
asset_rows = [
    ("People in file (LEX 8.85M + CCOM 1.40M)", 10252254, "10.25M consumers processed through the corrected matching pipeline; both populations carry dated tradelines through 2026 (data is CURRENT, not stale)"),
    ("Tradelines analyzed", 154300000, "135.9M LEX + 18.4M CCOM credit tradelines — a full-population corpus no plaintiff firm can replicate"),
    ("Credit inquiries", 117000000, "Impermissible-pull (FCRA §1681b) mining universe"),
    ("People matched to ≥1 legal claim signal", 5012704, "48.9% of the file"),
    ("Actionable today (live claim + contactable)", 1722586, "Matched, inside statute of limitations or no-SOL theory, not DNC-suppressed, intake-ready"),
    ("Contactability", None, "~100% name/address/email, 99% phone, consent flags at signup; ~800K DNC + ~200K email opt-outs already suppressed; CCOM book (1.4M) = currently-open customers = live outreach channel"),
]
for label, num, why in asset_rows:
    put(ws, r, 1, label, B)
    put(ws, r, 2, num, N, fmt=INT)
    put(ws, r, 5, why); r += 1

r += 1
put(ws, r, 1, "RECOVERY OPPORTUNITY — HEADLINE", H2); r += 1
r = header_row(ws, r, ["Opportunity layer", "Low", "Mid", "High", "What it is / why"])
headline = [
    ("A. Open settlements our base already sits in (next 12 months)",
     12_000_000, 35_000_000, 90_000_000,
     "12 live settlements totaling ~$769M in funds where measured cohorts of the file are class members (Capital One $425M, Navient $120M, Comcast $117.5M, Flagstar $31.5M, etc.). Range = member dollars captured at realistic claim-filing conversion; see 'Open Settlements' sheet.", GREEN),
    ("B. Bankruptcy §524 discharge violations — the flagship (12–36 mo)",
     12_000_000, 80_000_000, 290_000_000,
     "302,526 people verified by full-population SQL with a collection/charge-off still reporting a positive balance AFTER their bankruptcy filing. NO statute of limitations while reporting continues. $1K–$10K typical per-person contempt recovery. The single theory that converts the whole file into live claims. See '§524 Deep Dive' sheet.", GREEN),
    ("C. Originated consumer claims — FCRA / FDCPA / auto (12–36 mo)",
     25_000_000, 65_000_000, 150_000_000,
     "1.29M live-SOL signals + 1.99M live state-UDAP signals; 34.3M dispute-flagged tradelines (FCRA §1681i universe); 980K repossessions; Exeter mass-arbitration cohort of 235,993. Claims originated through the consented credit.com re-pull channel.", YELLOW),
    ("D. TCPA — phone-base settlement matching + repeat defendants (rolling)",
     2_000_000, 8_000_000, 25_000_000,
     "~9.9M consented phone numbers matchable against TCPA settlement dialer lists ($500–$1,500/call statutory). Near-term: Hy Cite $600–$1K/claimant, Athena Bitcoin, Citibank $29.5M. Pipeline: Credit One (403K), Synchrony (550K), Wells Fargo (392K) cohorts vs serial TCPA defendants.", YELLOW),
]
total = [0, 0, 0]
for label, lo, mid, hi, why, fill in headline:
    put(ws, r, 1, label, B, PatternFill("solid", fgColor=fill))
    put(ws, r, 2, lo, N, fmt=MONEYM); put(ws, r, 3, mid, B, fmt=MONEYM); put(ws, r, 4, hi, N, fmt=MONEYM)
    put(ws, r, 5, why)
    total = [t + v for t, v in zip(total, (lo, mid, hi))]; r += 1
put(ws, r, 1, "TOTAL CONSUMER RECOVERY OPPORTUNITY", H2)
for c, v in enumerate(total, 2):
    put(ws, r, c, v, B, PatternFill("solid", fgColor=GREEN), MONEYM)
put(ws, r, 5, "Gross dollars to credit.com members. Platform economics (rev-share / per-claim fee / counsel referral) layer on top — typical consumer-claims platforms monetize 15–30% of recovered value.", SM)
r += 1
put(ws, r, 1, "The 'Recovery by Case Type' sheet slices this same opportunity by statute instead of by timing layer; its totals differ modestly because open-settlement dollars span several statutes. Both views are funnel-adjusted estimates of the same pie, not additive.", SM)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
r += 2

put(ws, r, 1, "WHY CREDIT.COM IS UNIQUELY POSITIONED", H2); r += 1
for txt in [
    "1. PROOF: Credit data is the evidence. A tradeline showing a discharged debt still reporting, a disputed line never corrected, or a PHH mortgage opened 2007-2009 IS the class-membership proof — no other consumer platform holds it.",
    "2. CONSENT CHANNEL: Members already authorized credit monitoring. Settlement alerts and claim-filing assistance are a natural product surface — and consumer-initiated re-pulls are the legally clean intake path (FCRA permissible purpose).",
    "3. SCALE: Class settlements chronically under-claim (single-digit % claim rates). A 10M-person base with verified eligibility flips the economics: we find the member, pre-fill the claim, and capture dollars that otherwise revert.",
    "4. RECURRENCE: The same defendants (Capital One, Synchrony, Midland, Credit One, Santander...) settle repeatedly. The cohort index makes every future settlement a same-day matching exercise — this is a durable pipeline, not a one-off.",
]:
    put(ws, r, 1, txt); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5); r += 1

# ================================================================ 2. RECOVERY BY CASE TYPE
ws = sheet(wb, "Recovery by Case Type", [22, 22, 24, 14, 14, 26, 12, 12, 12, 52, 52])
title_block(ws, "Recovery Potential by Case Type",
            "Signals = claim-level matches in the file (one person can hold several). Recovery = funnel-adjusted dollars to members (see Funnel & Assumptions sheet), NOT statutory ceilings. Full docket lists on the 'Potential Cases' sheet.", 11)

def exn(label, pool):  # open-case count for a defendant label
    return len(pool.get(label, []))

ex524 = ("Originated contempt actions in the debtor's BK court — the template cases: Taggart v. Lorenzen (SCOTUS 2019, contempt standard); "
         "Haynes v. Chase (Bankr. S.D.N.Y., reporting/selling discharged debt as class contempt); Bruce v. Citigroup (2d Cir. 2023, §524 class survives arbitration push); "
         "In re Anderson v. Credit One (2d Cir.). Targets = the same furnishers our cohorts cover (Capital One, Credit One, TD, Wells Fargo Dealer, JPMCB Auto...).")
exFCRA = (f"{fdcpa_total_open} OPEN putative-class candidates (NOS 480) against file defendants right now: "
          f"LVNV ({exn('LVNV Funding', fdcpa_open)}), Midland ({exn('Midland Credit / Midland Funding', fdcpa_open)}), "
          f"Portfolio Recovery ({exn('Portfolio Recovery Associates', fdcpa_open)}), Credit One ({exn('Credit One Bank', fdcpa_open)}). "
          f"Freshest: {fmt_examples(fdcpa_open.get('Midland Credit / Midland Funding', []), 1)}.")
exFDCPA = (f"Same open NOS-480 pool — e.g. {fmt_examples(fdcpa_open.get('Portfolio Recovery Associates', []), 2)}. "
           "Every one is a certification-stage class our cohort members can join or mirror.")
exTCPA = (f"{tcpa_total_open} OPEN TCPA dockets (NOS 485) nationally; vs file defendants: "
          f"{fmt_examples(tcpa_open.get('Capital One', []), 1, 'court', 'docket_number')}; "
          f"{fmt_examples(tcpa_open.get('Credit One Bank', []), 1, 'court', 'docket_number')}. "
          "Plus paying windows: Hy Cite, Athena Bitcoin, Citibank $29.5M.")
exRESPA = "Munoz v. PHH (paying — $875/loan, deadline 8/11); Ocwen successor-liability theories on the same captive-reinsurance facts (52,681-person Ocwen cohort)."
exBreach = ("19 active data-breach/privacy MDLs tracked from PACER (1,453 member cases) — AT&T (MDL 3114), Change Healthcare (3108), "
            "Trans Union (3170), Snowflake (3126), MOVEit (3083)... each one becomes a claims window our base matches into; "
            "paying now: Comcast, Flagstar, LastPass, Fidelity.")
exAuto = (f"{exn('Santander Consumer USA', fdcpa_open)} open Santander class candidates + Exeter mass-arb (rolling) — e.g. "
          f"{fmt_examples(fdcpa_open.get('Santander Consumer USA', []), 1)}; Westlake ({exn('Westlake Financial / Westlake Portfolio', fdcpa_open)} open).")
exStudent = "CFPB v. Navient redress (paying now); state AG servicing suits vs MOHELA/Aidvantage as they mature."
exUDAP = "Ford v. Genesis/Concora (MD, open to 6/29) is the template — creditor + state statute; payday/high-cost lender tradelines map to state UDAP regimes."

r = 4
r = header_row(ws, r, ["Case type", "Statute / theory", "Per-claimant damages", "Claim signals in file",
                       "Est. eligible people", "SOL posture", "Recovery LOW", "Recovery MID", "Recovery HIGH",
                       "Why it works with this data", "Potential cases (live dockets / templates)"])
case_rows = [
    ("Bankruptcy §524 Discharge Violation", "11 U.S.C. §524 / contempt (Taggart v. Lorenzen)",
     "$1,000–$10,000+ typical settlements (actual + emotional distress + fees; punitive for willful)",
     3865163, 302526, "NO SOL — violation is ongoing while the discharged debt keeps reporting",
     12_000_000, 80_000_000, 290_000_000,
     "302,526 people verified by SQL: collection/charge-off with positive balance reported AFTER their BK filing date (436,003 total bankruptcy filers in file). The tradeline itself is the violation evidence. Funnel: consented re-pull confirms ~40-60% still reporting → 10-20% retain counsel → 12K-36K claimants.", ex524, GREEN),
    ("FCRA — credit reporting", "15 U.S.C. §1681n/o (willful: $100–$1,000 statutory + punitive; negligent: actual)",
     "$250–$1,000 realistic per settled claim",
     19468010, 700000, "2 yrs from discovery / 5 yrs from violation — 1.29M signals currently live",
     8_000_000, 25_000_000, 60_000_000,
     "34.3M tradelines carry the dispute flag (disp=1) — the §1681i dispute-ignored universe. Full-population per-furnisher analytics (dispute-ignored rate, re-aging fingerprints, double-sold debt) also sell as Systemic Violation Reports to class counsel at $15-50K each.", exFCRA, GREEN),
    ("FDCPA — debt collection", "15 U.S.C. §1692k ($1,000 statutory + actual; class cap 1% net worth)",
     "$300–$1,000",
     4493625, 300000, "1 yr — short, but collection activity refreshes continuously; live subset only",
     5_000_000, 15_000_000, 30_000_000,
     "Collector tradelines (Midland 241K, NCA, LVNV, Portfolio Recovery...) identify the relationship; ongoing reporting/collection inside 12 months creates fresh claims every month. 31,305-docket PACER FDCPA/FCRA index proves these defendants settle constantly.", exFDCPA, None),
    ("TCPA — calls/texts", "47 U.S.C. §227(b)(3): $500/call, $1,500 willful; FL FTSA parallel",
     "$500–$1,500 individual; $20–$100 class pro-rata; Hy Cite paying $600–$1,000",
     None, None, "4 yrs — sized via phone-base matching, NOT tradelines (call receipt isn't in credit data)",
     2_000_000, 8_000_000, 25_000_000,
     "99% phone coverage with consent = administrator dialer-list matching at scale. 10,288-docket national TCPA index built from PACER. Repeat collection-call defendants overlap our biggest cohorts: Credit One 403K, Synchrony 550K, Wells Fargo 392K people.", exTCPA, None),
    ("RESPA — mortgage", "12 U.S.C. §2607 (§8 kickbacks); servicing-transfer violations",
     "PHH: $875/loan FIXED, no pro-rata cut",
     448297, 5804, "Open claim window (PHH deadline 2026-08-11)",
     1_300_000, 2_500_000, 5_100_000,
     "Cleanest Tier-1 match in the sweep: 5,804 people (4,180 LEX + 1,624 CCOM) hold a PHH mortgage opened 2007-2009 — the exact class window. $5.1M claim ceiling at $875/loan. High = full ceiling; mid = 50% targeted-outreach conversion.", exRESPA, GREEN),
    ("Data breach settlements", "Negligence / state privacy statutes (settled classes)",
     "$50–$599 no-proof tiers; up to $25K documented; +$100 CA",
     53538, 1300000, "Settlement claim windows (Jun–Sep 2026); membership = PII in breach",
     3_000_000, 8_000_000, 20_000_000,
     "~$202M in open breach funds: Comcast $117.5M (~31.6M class ≈ 12% of any US base → ~1.2M of our file), Flagstar $31.5M (38,418 measured tradeline holders), LastPass $24.45M, Fidelity. Email base + attestation; breach classes under-claim at 1-3% — assisted filing captures reverting dollars.", exBreach, None),
    ("Auto lending / repossession (incl. mass arbitration)", "State UDAP, UCC Art. 9, arb-clause mass filing",
     "$1,500–$7,500 typical mass-arb resolution (fee-exposure leverage)",
     1622484, 980000, "State UDAP 3–5 yrs; mass-arb rolling, no deadline",
     6_000_000, 18_000_000, 45_000_000,
     "980K repossessions in full population. Exeter mass-arb is LIVE and rolling: 235,993 people with Exeter tradelines (Tier 1). Santander 557K, Westlake 201K, Credit Acceptance cohorts behind it. 2-5% participation on Exeter alone = 4.7K-11.8K claimants.", exAuto, None),
    ("Student loan servicing", "CFPB orders, state servicing statutes",
     "CFPB-determined redress (Navient checks mailing now)",
     1725070, 451340, "Navient redress PAYING since 2026-02-13, automatic",
     1_000_000, 3_000_000, 8_000_000,
     "451,340 Navient-tradeline people identified Tier-1. $120M redress paying automatically — the play is address-currency outreach so checks land (plus MOHELA/Aidvantage pipeline as state cases mature).", exStudent, None),
    ("State UDAP / payday", "State unfair-practices acts (FDUTPA etc.)",
     "Varies; actual + fees, some statutory",
     823569, 250000, "3–5 yr state SOLs — 1.99M signals live under state law",
     2_000_000, 6_000_000, 15_000_000,
     "High-cost lender and payday tradelines in states with strong UDAP regimes; Genesis/Milestone Maryland settlement (open to 6/29) is the template — creditor + state = Tier-1 class membership.", exUDAP, None),
]
tot = [0, 0, 0]
for row in case_rows:
    (ct, stat, dmg, sigs, ppl, sol, lo, mid, hi, why, ex, fill) = row
    put(ws, r, 1, ct, B, PatternFill("solid", fgColor=fill) if fill else None)
    put(ws, r, 2, stat); put(ws, r, 3, dmg)
    put(ws, r, 4, sigs, N, fmt=INT); put(ws, r, 5, ppl, N, fmt=INT)
    put(ws, r, 6, sol)
    put(ws, r, 7, lo, N, fmt=MONEYM); put(ws, r, 8, mid, B, fmt=MONEYM); put(ws, r, 9, hi, N, fmt=MONEYM)
    put(ws, r, 10, why); put(ws, r, 11, ex)
    tot = [t + v for t, v in zip(tot, (lo, mid, hi))]; r += 1
put(ws, r, 1, "TOTAL", H2)
for c, v in enumerate(tot, 7): put(ws, r, c, v, B, PatternFill("solid", fgColor=GREEN), MONEYM)
r += 2
put(ws, r, 1, "Notes: 'Claim signals' count claim-level matches (one person can carry several; FCRA signals count disputed lines). 'Est. eligible people' is the de-duplicated person-level cohort where measured (SQL/index), otherwise a labeled estimate. TCPA is intentionally NOT derived from tradelines — owning a phone is not a TCPA claim; it is sized via settlement dialer-matching and repeat-defendant cohorts. Docket citations in the last column are live PACER records (status open as of 2026-06-08/09 index pull); full lists on the 'Potential Cases' sheet.", SM)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
ws.freeze_panes = "A5"

# ================================================================ 2b. POTENTIAL CASES
ws = sheet(wb, "Potential Cases", [16, 30, 13, 36, 13, 16, 60])
title_block(ws, "Potential Cases — Live Dockets Backing Each Opportunity",
            "Mined from our PACER enumeration (31,305 FDCPA/FCRA + 10,288 TCPA dockets + 19 breach MDLs; status as of the 2026-06-08/09 index pull). 'Open class candidates' = open putative class actions (or cases that ran >1.5 yrs) naming the defendant.", 7)
r = 4

put(ws, r, 1, f"FDCPA / FCRA — {fdcpa_total_open:,} OPEN CLASS CANDIDATES NATIONALLY (NOS 480)", H2); r += 1
r = header_row(ws, r, ["Case type", "Defendant", "Open class candidates", "Most recent open filings (court · docket · filed)",
                       "Cohort in file", "Status", "What joining/mirroring looks like"])
fdcpa_notes = {
    "LVNV Funding": "Largest debt-buyer docket footprint in our index (1,063 total dockets). Top target for the per-furnisher Systemic Violation Report (dispute-ignored rate, re-aging).",
    "Midland Credit / Midland Funding": "Perennial FDCPA defendant; Leedeman (CA) already paying our members automatically — these are the next windows.",
    "Portfolio Recovery Associates": "CFPB repeat offender (2015 + 2023 orders) — pattern evidence strengthens every class.",
    "Santander Consumer USA": "Subprime auto: repo notices, deficiency calculation, fee theories; multistate AG history.",
    "Credit One Bank": "Subprime card fees + collection practices; cohort 403,156.",
    "Westlake Financial / Westlake Portfolio": "Klare just paid (closed 5/18/26); these are the successor actions.",
    "Jefferson Capital Systems": "Debt buyer; chain-of-title attacks viable with our double-sold-debt analytics.",
    "Ally Financial": "Auto servicing and repo notice theories.",
    "Transworld Systems": "High-volume collector; student-loan NCSLT connection.",
    "I.C. System": "High-volume medical/utility collector — dispute-handling theories.",
    "Credit Acceptance": "Subprime auto; Mass. AG / CFPB history.",
    "Exeter Finance": "Class track parallel to our rolling mass-arb program.",
    "First Premier Bank": "Fee-harvester card theories.",
    "OneMain Financial": "Add-on products / insurance packing theories.",
    "Capital One": "Largest cohort in file (1,712,966) — any certification is instantly material.",
    "National Credit Adjusters": "Blackburn (VA) paying now; these are the other jurisdictions.",
}
for label, _pat in FDCPA_DEFS:
    cases = fdcpa_open.get(label, [])
    if not cases:
        continue
    coh = next((v for k, v in COHORTS.items() if k in label.lower()), None)
    put(ws, r, 1, "FDCPA/FCRA")
    put(ws, r, 2, label, B)
    put(ws, r, 3, len(cases), N, fmt=INT, align=WRAPC)
    put(ws, r, 4, fmt_examples(cases, 2))
    put(ws, r, 5, coh, N, fmt=INT)
    put(ws, r, 6, "OPEN", N, PatternFill("solid", fgColor=GREEN), align=WRAPC)
    put(ws, r, 7, fdcpa_notes.get(label, "")); r += 1

r += 1
put(ws, r, 1, f"TCPA — {tcpa_total_open:,} OPEN DOCKETS NATIONALLY (NOS 485)", H2); r += 1
r = header_row(ws, r, ["Case type", "Defendant", "Open TCPA dockets", "Most recent open filings (court · docket · filed)",
                       "Cohort in file", "Status", "What joining/mirroring looks like"])
tcpa_notes = {
    "Capital One": "Collection-call classes against our biggest cohort; Capital One hits 1,653 of the top 2,000 scored people.",
    "Credit One Bank": "Serial autodialer defendant — the canonical credit-cohort TCPA target.",
    "Westlake Services / Portfolio Mgmt": "Collection calls on auto accounts; cohort 200,689.",
    "Portfolio Recovery Associates": "Collection-call TCPA on purchased debt.",
    "Wells Fargo": "Collection robocall history (prior nine-figure TCPA settlements).",
    "Synchrony Bank": "Store-card collection calls; cohort 549,694.",
    "Santander Consumer USA": "Auto collection calls.",
    "Bank of America": "Prior $32M TCPA settlement precedent (Rose).",
    "Citibank": "Citibank $29.5M TCPA settlement already seeded in our catalog — docket-verify per-claimant terms.",
    "Ally Financial": "Auto-account collection calls.",
}
for label, _pat in TCPA_DEFS:
    cases = tcpa_open.get(label, [])
    if not cases:
        continue
    coh = next((v for k, v in COHORTS.items() if k in label.lower()), None)
    put(ws, r, 1, "TCPA")
    put(ws, r, 2, label, B)
    put(ws, r, 3, len(cases), N, fmt=INT, align=WRAPC)
    put(ws, r, 4, fmt_examples(cases, 2, "court", "docket_number"))
    put(ws, r, 5, coh, N, fmt=INT)
    put(ws, r, 6, "OPEN", N, PatternFill("solid", fgColor=GREEN), align=WRAPC)
    put(ws, r, 7, tcpa_notes.get(label, "")); r += 1

r += 1
put(ws, r, 1, "DATA BREACH / PRIVACY — 19 ACTIVE MDLs (1,453 MEMBER CASES)", H2); r += 1
r = header_row(ws, r, ["Case type", "MDL", "Member cases", "Open", "", "Status", "Why it matters to the base"])
mdl_notes = {
    "3114": "AT&T 2024 breach (~110M records) — telecom overlap with a consumer base is near-universal.",
    "3108": "Change Healthcare — up to 1 in 3 Americans' health data; population-overlap play like Comcast.",
    "3170": "Trans Union — a credit bureau breach maps 1:1 onto a credit-monitoring membership.",
    "3126": "Snowflake (Ticketmaster/Santander Bank et al.) — Santander overlap touches our 557K cohort.",
    "3083": "MOVEit — largest member-case MDL; dozens of sub-settlements will open claim windows for years.",
    "3098": "23andMe — bankruptcy complicates but claims process survives.",
    "3073": "T-Mobile 2022 — follow-on to the $350M 2021 settlement template.",
    "3127": "Evolve Bank — fintech/BaaS users (Affirm, Mercury) overlap young-credit demographics.",
    "3149": "PowerSchool — parents/students; broad household overlap.",
}
for (num, name), (tot_c, open_c) in sorted(mdl_agg.items(), key=lambda x: -x[1][0])[:12]:
    put(ws, r, 1, "Data breach")
    put(ws, r, 2, f"MDL {num} — {name}", B)
    put(ws, r, 3, tot_c, N, fmt=INT, align=WRAPC)
    put(ws, r, 4, open_c, N, fmt=INT, align=WRAPC)
    put(ws, r, 6, "PRE-SETTLEMENT", N, PatternFill("solid", fgColor=YELLOW), align=WRAPC)
    put(ws, r, 7, mdl_notes.get(num, "Each MDL resolution opens a claims window the base matches into.")); r += 1

r += 1
put(ws, r, 1, "BANKRUPTCY §524 — ORIGINATED ACTIONS (TEMPLATE PRECEDENTS)", H2); r += 1
r = header_row(ws, r, ["Case type", "Precedent / matter", "", "What it established", "", "Status", "How we use it"])
rows524 = [
    ("Taggart v. Lorenzen, 588 U.S. 554 (2019)", "SCOTUS contempt standard for §524: no fair ground of doubt the conduct was unlawful.",
     "Reporting a positive balance on a discharged debt, after notice, clears the willfulness bar — the demand letter creates the record."),
    ("Haynes v. Chase Bank (Bankr. S.D.N.Y.)", "Selling/reporting discharged debt prosecuted as CLASS-WIDE contempt; drove major-bank settlements.",
     "The structural template for running our 302,526-person cohort as defendant-by-defendant class contempt rather than one-offs."),
    ("Bruce v. Citigroup, 75 F.4th 297 (2d Cir. 2023)", "§524 class action survives the arbitration-clause defense — bankruptcy court keeps it.",
     "Removes the defendants' favorite exit door (the same one that kills consumer class actions elsewhere)."),
    ("In re Anderson v. Credit One (2d Cir. 2018)", "Credit One could not compel arbitration of discharge-violation claims; refusal to correct tradelines actionable.",
     "Credit One is simultaneously a top §524 target AND a 403,156-person cohort in our file."),
]
for cite, held, use in rows524:
    put(ws, r, 1, "§524 discharge")
    put(ws, r, 2, cite, B)
    put(ws, r, 4, held)
    put(ws, r, 6, "TEMPLATE", N, PatternFill("solid", fgColor=SUB_FILL.fgColor.rgb), align=WRAPC)
    put(ws, r, 7, use); r += 1

r += 1
put(ws, r, 1, "Sources: data/pacer-cases/_candidates.json (3,218 class candidates from 31,305 FDCPA/FCRA dockets), _tcpa_index.json (10,288 NOS-485 dockets), _breach_mdl_index.json (JPML member-case enumeration). Counts reflect docket status at index pull; re-run tools/pacer-enumerate.py to refresh. Settled/paying matters live on the 'Open Settlements' sheet — this sheet is the FORWARD pipeline.", SM)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
ws.freeze_panes = "A5"

# ================================================================ 3. OPEN SETTLEMENTS
ws = sheet(wb, "Open Settlements", [26, 30, 14, 13, 20, 16, 10, 14, 50])
title_block(ws, "Live Settlements Touching the File (June 2026)",
            "Bucket A = already settled, members get paid with no/minimal action (product = awareness + address currency). Bucket B = open claim window (product = assisted claim filing). Cohorts measured against corrected v2 data.", 9)
r = 4
put(ws, r, 1, "BUCKET A — FRICTIONLESS (automatic payment, ~2.7M+ of the file)", H2); r += 1
r = header_row(ws, r, ["Defendant", "Settlement", "Fund", "Status / deadline", "Per-claimant", "Cohort in file", "Tier", "Est. member $ captured", "Play"])
bucketA = [
    ("Capital One", "360 Savings interest-rate litigation", "$425M", "Auto-pay ~2026-07-21", "Pro-rata lost interest", 1712966, 2, 8_000_000,
     "Deposit accounts are invisible to credit data, so the 1.71M Cap One cohort is the outreach ceiling. Awareness + address-verification campaign; even a small slice of a $425M automatic fund landing correctly is real member value."),
    ("Navient", "CFPB v. Navient — $120M redress", "$120M", "Checks mailing since 2026-02-13", "CFPB-determined", 451340, 1, 5_000_000,
     "Tier-1: servicer tradeline directly identifies the cohort. Address-currency campaign so checks land; zero filing cost."),
    ("Bank of America", "7-Eleven FCTI ATM fees", "$2.25M", "Former customers file by 2026-06-29", "Fee refunds", 333611, 2, 150_000,
     "Current customers auto-credited; simple attestation claim for former ones."),
    ("Midland Credit Mgmt", "Leedeman v. MCM (CA) — automatic", "see workbook", "Automatic, optional e-pay election", "Varies", 22000, 1, 500_000,
     "241,303 national MCM cohort; ~9% CA ≈ 22K paid without acting. Natural credit.com story: the collection relief shows on the report."),
    ("National Credit Adjusters", "Blackburn v. NCA (VA) — automatic", "debt cancel + deletion", "Automatic", "Debt cancellation + tradeline deletion", None, 1, None,
     "Relief appears on the credit report itself — credit.com can show members their score improving."),
]
for d, s, f, st, pc, coh, tier, est, play in bucketA:
    put(ws, r, 1, d, B); put(ws, r, 2, s); put(ws, r, 3, f, align=WRAPC); put(ws, r, 4, st)
    put(ws, r, 5, pc); put(ws, r, 6, coh, N, fmt=INT)
    put(ws, r, 7, f"Tier {tier}", N, PatternFill("solid", fgColor=GREEN if tier == 1 else YELLOW), align=WRAPC)
    put(ws, r, 8, est, N, fmt=MONEY); put(ws, r, 9, play); r += 1

r += 1
put(ws, r, 1, "BUCKET B — OPEN CLAIM WINDOWS (we run the filing workflow)", H2); r += 1
r = header_row(ws, r, ["Defendant", "Settlement", "Fund", "Deadline", "Per-claimant", "Cohort in file", "Tier", "Est. member $ captured", "Play"])
bucketB = [
    ("PHH Mortgage", "Munoz — RESPA captive-reinsurance kickbacks", "Fixed per-loan", "2026-08-11", "$875/loan FIXED", 5804, 1, 2_500_000,
     "TOP ITEM. Mortgage tradeline + 2007-2009 open date = the class definition. $5.1M ceiling; 50% conversion on targeted Tier-1 outreach is achievable because we pre-identify exact eligibility."),
    ("Flagstar Bank", "2021 data breaches", "$31.5M", "2026-08-11", "~$60 base to $599; $25K documented; +$100 CA", 38418, 2, 1_500_000,
     "38,418 measured Flagstar tradeline holders vs 2.19M class — meaningful Claim-ID overlap. Workflow includes administrator lookup for the notice Claim ID."),
    ("Comcast / Xfinity", "2023 Citrix Bleed breach (Kroll)", "$117.5M", "2026-08-14 (verify 9/14 ext.)", "$50 base; $10K documented", 1200000, 2, 4_000_000,
     "Creditor match negligible — this is a population-overlap play: ~31.6M class ≈ 12% of any US adult base → ~1.2M of the file plausibly eligible. No-proof $50 tier = email campaign + attestation."),
    ("Hy Cite / Royal Prestige", "TCPA prerecorded calls (non-customers)", "$4.75M", "2026-07-08", "$600–$1,000 est.", None, 2, 750_000,
     "Highest per-claimant cash open right now. Phone numbers on file enable administrator dialer-list matching; pure attestation claim."),
    ("Hyundai / Kia", "Immobilizer theft (multistate)", "$4.5M + $9M", "2027-03-31", "$375–$4,500 by tier", 129767, 2, 1_500_000,
     "59,915 Hyundai + 69,852 Kia finance relationships (gross). 2011-2022 turn-key models + documented theft narrows it, but the 21-month window makes this a durable pipeline item."),
    ("LastPass", "2022 security incident", "$24.45M", "2026-07-02", "$25–$10,400 by tier (+crypto pool)", None, 2, 300_000,
     "No credit-data signal; email-base play to the 10M list, notice-gated."),
    ("Genesis / Concora (Milestone)", "Ford — MD unlicensed lending", "see workbook", "2026-06-29", "Varies", None, 1, 250_000,
     "Tier-1: Genesis/Milestone tradeline + MD address. Mailed PIN required — tight window."),
    ("Athena Bitcoin", "Jackson — TCPA/FTSA texts", "$4.5M", "2026-06-30", "Share of fund", None, 2, 150_000,
     "Notice-gated (Class Member ID); phone match via administrator."),
    ("Amazon Prime", "FTC dark-patterns redress", "FTC order", "2026-07-27", "Up to $51", None, 2, 500_000,
     "Universal no-proof claim — pure email-base volume play."),
    ("Fidelity", "Aug 2024 breach", "$2.5M", "2026-07-27", "~$100 base (+$50 CA)", None, 2, 25_000,
     "77K class — included for completeness."),
    ("Lakeview Loan Servicing", "Oct 2021 breach", "$26M", "2026-06-22", "Pro rata + $5K documented", 200, 2, None,
     "DEPRIORITIZED: ~200 matches are name collisions; Lakeview subservices via LoanCare so it never reports. Attestation-only, tiny window."),
]
for d, s, f, dl, pc, coh, tier, est, play in bucketB:
    put(ws, r, 1, d, B); put(ws, r, 2, s); put(ws, r, 3, f, align=WRAPC); put(ws, r, 4, dl, align=WRAPC)
    put(ws, r, 5, pc); put(ws, r, 6, coh, N, fmt=INT)
    put(ws, r, 7, f"Tier {tier}", N, PatternFill("solid", fgColor=GREEN if tier == 1 else YELLOW), align=WRAPC)
    put(ws, r, 8, est, N, fmt=MONEY); put(ws, r, 9, play); r += 1

r += 1
put(ws, r, 1, "BUCKET C — MASS ARBITRATION (rolling, no deadline)", H2); r += 1
r = header_row(ws, r, ["Defendant", "Program", "Fund", "Deadline", "Per-claimant", "Cohort in file", "Tier", "Est. member $ captured", "Play"])
put(ws, r, 1, "Exeter Finance", B); put(ws, r, 2, "Mass arbitration — undisclosed processing/convenience fees")
put(ws, r, 3, "Individual awards", align=WRAPC); put(ws, r, 4, "Rolling", align=WRAPC)
put(ws, r, 5, "$1,500–$7,500 typical resolution range"); put(ws, r, 6, 235993, N, fmt=INT)
put(ws, r, 7, "Tier 1", N, PatternFill("solid", fgColor=GREEN), align=WRAPC)
put(ws, r, 8, 15_000_000, N, fmt=MONEY)
put(ws, r, 9, "Exeter auto tradeline = direct Tier-1 relationship proof. 2-5% participation = 4.7K-11.8K claimants; arbitration-fee exposure drives defendant settlement. No deadline — durable program.")
r += 2
put(ws, r, 1, "Tier 1 = our data fields directly satisfy the class definition (creditor + product + date window). Tier 2 = data establishes the plausible pool; the gating fact (call received, breach notice, theft event) needs claimant attestation or administrator match. 'Est. member $ captured' is the funnel-adjusted mid case, not the fund size.", SM)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
ws.freeze_panes = "A5"

# ================================================================ 4. DEFENDANT COHORTS
ws = sheet(wb, "Defendant Cohorts", [26, 16, 22, 30, 60])
title_block(ws, "Top Defendant Cohorts in the File",
            "People with a tradeline/collection relationship to each defendant — production sharded index (rederive v2, LEX + CCOM). These are the audiences for every current AND future settlement by that defendant.", 5)
r = 4
r = header_row(ws, r, ["Defendant", "People in file", "Litigation profile (PACER)", "Active opportunity now", "Why this cohort matters"])
cohorts = [
    ("Capital One", 1712966, "Perennial defendant — all consumer case types", "360 Savings $425M auto-paying Jul 2026",
     "Largest cohort in the file. Every future Cap One settlement (fees, reporting, TCPA) is a same-day match against 1.7M pre-identified members."),
    ("Santander Consumer USA", 556939, "Subprime auto; multistate AG settlements; repo practices", "Structural — next repo/fee action",
     "Subprime auto book with documented regulatory history; repossession + deficiency theories recur."),
    ("Synchrony Bank", 549694, "Serial TCPA collection-call defendant", "TCPA pipeline",
     "Store-card collections at scale → recurring autodialer settlements; 550K relationship holders to match against dialer lists."),
    ("Navient", 451340, "CFPB order + state servicing cases", "$120M redress PAYING NOW (Tier 1)",
     "Direct Tier-1 identification of redress-eligible borrowers."),
    ("Credit One Bank", 403156, "TCPA + FDCPA recurring", "TCPA/FDCPA pipeline",
     "Subprime card issuer with constant collection-practice litigation."),
    ("Wells Fargo", 392187, "TCPA collection robocalls; fee cases", "Structural",
     "Repeat settler across multiple consumer theories."),
    ("Bank of America", 333611, "Fees / TCPA history", "ATM fee settlement (file by 6/29)",
     "Bucket-A automatic payment for current customers."),
    ("Midland Credit Management", 241303, "FDCPA perennial (debt buyer)", "Leedeman CA automatic",
     "Debt-buyer FDCPA claims regenerate continuously; CA members being paid now without acting."),
    ("Exeter Finance", 235993, "Subprime auto fees", "LIVE mass arbitration (rolling)",
     "Tier-1 mass-arb audience — the only program here with no deadline."),
    ("Westlake Financial", 200689, "FDCPA / repossession", "Klare settled 5/18/26; next action",
     "Recurring defendant; cohort ready for the next window."),
    ("PHH / Ocwen", 63427, "RESPA captive reinsurance", "PHH $875/loan, deadline 8/11 (Tier 1)",
     "5,804 in the exact 2007-09 class window now; 63K total mortgage relationships for successor theories."),
    ("Flagstar Bank", 38418, "2021 breaches", "$31.5M fund, deadline 8/11",
     "Measured tradeline overlap with a 2.19M-person class."),
    ("Hyundai / Kia Finance", 129767, "Immobilizer theft MDL + multistate", "$13.5M funds to 2027-03-31",
     "Finance tradelines identify owners/lessees of affected model years."),
]
for d, n, lit, now, why in cohorts:
    put(ws, r, 1, d, B); put(ws, r, 2, n, N, fmt=INT); put(ws, r, 3, lit); put(ws, r, 4, now); put(ws, r, 5, why); r += 1
r += 1
put(ws, r, 1, "Litigation evidence base: 31,305 FDCPA/FCRA dockets + 10,288 national TCPA dockets + 1,453 data-breach MDL member cases enumerated from PACER against these defendants (41 canonical creditor/collector entities; 1,276 open cases). Cross-population (LEX vs CCOM) dedup pending — same person can appear in both books.", SM)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
ws.freeze_panes = "A5"

# ================================================================ 5. §524 DEEP DIVE
ws = sheet(wb, "§524 Deep Dive", [40, 18, 80])
title_block(ws, "Bankruptcy Discharge Violations — The Flagship Opportunity",
            "Why this is the single most valuable theory in the file", 3)
r = 4
rows524 = [
    ("THE THEORY", None,
     "A creditor/collector that keeps reporting a positive balance on a debt discharged in bankruptcy is attempting to collect a discharged debt — a violation of the 11 U.S.C. §524 discharge injunction, enforced through contempt (Taggart v. Lorenzen, 588 U.S. 554 (2019)). Courts award actual damages, emotional distress, attorney fees, and punitive sanctions for willful violations."),
    ("WHY IT BEATS THE SOL PROBLEM", None,
     "FDCPA gives 1 year, FCRA 2-5, TCPA 4. §524 has NO statute of limitations — the violation is ONGOING for as long as the false balance keeps reporting. This is the one theory that converts historical credit data into live, filable claims today, every day."),
    ("VERIFIED COHORT", 302526,
     "Full-population SQL on the 154M-tradeline corpus: people showing a Collection or Charge-Off tradeline with a POSITIVE BALANCE reported AFTER their bankruptcy filing date. Not modeled — counted."),
    ("Total bankruptcy filers in file", 436003,
     "512,939 BK public records; 78% filed before their credit.com signup (they came to credit.com to rebuild — and the file shows who is still being wrongly reported)."),
    ("Recent filers (BK 2020+)", 31923,
     "Freshest discharge dates = cleanest damages narrative and most active reporting violations."),
    ("PER-CLAIMANT VALUE", None,
     "Reported §524 outcomes commonly run $1,000–$10,000 per debtor in settlement (fee-shifting drives defendant economics); willful/egregious cases reach $25K+. Defense cost per contested contempt motion exceeds most settlements — defendants resolve."),
    ("THE FUNNEL", None,
     "302,526 cohort → consented re-pull confirms line still reporting (40–60%) → retains counsel (10–20%) → 12,000–36,000 claimants → at $1K–$8K each = $12M low / $80M mid / $290M high gross recovery."),
    ("WHY CREDIT.COM", None,
     "These are people who came to credit.com AFTER bankruptcy to rebuild credit. Telling them 'a discharged debt is still being held against your score — and that's illegal' is simultaneously the product promise, the legal claim, and the consented intake event. The re-pull they authorize IS the evidence."),
    ("EXECUTION STATUS", None,
     "Cohort built and indexed (DischargeViolation signals live in production matching). Next: pre-petition open-date screen finalized (done for CCOM, LEX re-run queued), CourtListener discharge-date join for the damages model, and counsel-network routing."),
]
for label, num, txt in rows524:
    put(ws, r, 1, label, B, SUB_FILL if num is None else None)
    put(ws, r, 2, num, B, fmt=INT)
    put(ws, r, 3, txt); r += 1
ws.freeze_panes = "A4"

# ================================================================ 6. FUNNEL & ASSUMPTIONS
ws = sheet(wb, "Funnel & Assumptions", [34, 20, 80])
title_block(ws, "How We Get From Cohorts to Dollars — Honest Math",
            "Every recovery figure in this workbook is funnel-adjusted. We size on screened conversion, never on raw cohort ceilings.", 3)
r = 4
r = header_row(ws, r, ["Funnel stage", "Rate applied", "Basis"])
funnel = [
    ("1. Cohort ceiling", "100%", "People with the qualifying creditor relationship (measured by index/SQL, not modeled)."),
    ("2. Contactable", "~85%", "99% phone / ~100% email coverage, minus ~800K DNC and ~200K email opt-outs (already flagged in file). CCOM book (1.4M open customers) converts best."),
    ("3. Class-window fit", "30–70%", "Date window / state / product narrowing per class definition. Tier 1 = provable from our fields; Tier 2 = plausible pool."),
    ("4. Gating event (Tier 2 only)", "5–40%", "Received the call / breach notice / theft event — attested by claimant or matched by administrator."),
    ("5. Claim filed / retained", "Tier 1 targeted: 25–50% · Tier 2 assisted: 5–15% · blind email: 1–3%",
     "Industry baseline class-action claim rates are single-digit; pre-identified eligibility + pre-filled forms is precisely what moves them. Mass-arb/§524 retention modeled at 10–20% of confirmed-eligible."),
    ("6. Payment lands", "95%+", "Address verification is itself part of the product for automatic (Bucket A) settlements."),
]
for s, rate, basis in funnel:
    put(ws, r, 1, s, B); put(ws, r, 2, rate); put(ws, r, 3, basis); r += 1
r += 1
put(ws, r, 1, "COMPLIANCE GUARDRAILS (built into every estimate)", H2); r += 1
for txt in [
    "FCRA §1681b permissible purpose: purchased credit data is used for AGGREGATE analytics and sizing only. Individual claims run through consumer-initiated, consented re-pulls (the credit.com webhook channel) — the legally clean intake path.",
    "ABA Rule 7.3 / barratry: credit.com surfaces settlement awareness and claim-filing tools to its own members; counsel relationships are member-initiated referrals, not data-driven solicitation.",
    "Telemarketing posture: outreach designed email/in-app first. DNC (~800K) and email opt-outs (~200K) suppressed at the data layer; PGX consent-order context (10-yr telemarketing-credit-repair ban) respected in channel design.",
    "Numbers hygiene: signal counts vs person counts are labeled throughout; cross-population dedup (LEX/CCOM) pending; every cohort is a measured ceiling, every dollar figure is funnel-adjusted.",
]:
    put(ws, r, 1, txt); ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3); r += 1

# ================================================================ 7. DATA & METHODOLOGY
ws = sheet(wb, "Data & Methodology", [36, 96])
title_block(ws, "Data Assets and How the Numbers Were Built", "All figures dated 2026-06-10, corrected v2 derivation", 2)
r = 4
meth = [
    ("Source data", "LEX book: 8.85M identities, 135.9M tradelines, ~117M inquiries, 1.4M public records, dated through 2026. CCOM book: 1.4M currently-open credit.com customers, 18.4M tradelines (DateOpened/BalanceDate/FilingDate), 84,670 bankruptcy records. Contact: ~100% name/address/email, 99% phone, signup consent flags."),
    ("Matching pipeline", "Full-population rederivation (June 2026): every person scored against a canonical defendant library + statute-of-limitations engine per claim. 10,252,254 processed → 5,012,704 matched → 1,722,586 actionable (live + contactable + not DNC). Fabricated signal types from the earlier pipeline (TCPA-from-tradeline, breach-from-ownership) were REMOVED — what remains is defensible."),
    ("SOL engine", "Each claim signal carries live / live-state-UDAP / discharge-ongoing / time-barred status from actual tradeline dates: 1.29M live + 1.99M live-UDAP + 3.87M discharge-ongoing vs 25.4M honestly marked time-barred."),
    ("Litigation index", "PACER enumeration against the top creditor/debt-buyer entities: 31,305 FDCPA/FCRA dockets, 10,288 national TCPA dockets, 19 data-breach/privacy MDLs (1,453 member cases), 3,218 class-action candidates — joined to a per-defendant settlement-intake workbook (administrators, claim windows, URLs, required proof)."),
    ("Settlement catalog", "Wave-1 sweep of administrator sites + Top Class Actions / ClassAction.org / openclassactions (June 2026), verified against administrator pages where available; cohort sizes measured by direct creditor-pattern SQL over the full tradeline corpus (all spelling variants — e.g. 7 PHH spellings, 8 Ocwen spellings)."),
    ("What 'Tier 1' means", "Our fields alone satisfy the class definition (creditor + product + date window) — e.g. PHH mortgage opened 2007-2009, Navient student-loan tradeline, Exeter auto loan. Tier 2 = we establish the plausible pool; the gating fact needs attestation or administrator match."),
    ("Known limitations", "(1) LEX/CCOM cross-population overlap not yet deduplicated. (2) LEX §524 signals await re-run with the stricter pre-petition screen already applied to CCOM (the 302,526 SQL cohort already applies it). (3) Tier-2 cohorts are ceilings on the plausible pool, not confirmed class membership. (4) Recovery ranges are funnel-adjusted estimates, not guarantees; per-claimant values cite statutory text and observed settlement outcomes."),
    ("Regeneration", "tools/build-creditcom-recovery-xlsx.py rebuilds this workbook; underlying artifacts: data/settlements/open-settlements-2026-06.json, data/settlements/recovery-sizing-2026-06-10.md, settlement_cohort_sizes.json, production casepeople index, docs/Field-Population-Walkthrough-2026-06-10.md."),
]
for label, txt in meth:
    put(ws, r, 1, label, B, SUB_FILL); put(ws, r, 2, txt); r += 1

wb.save(OUT)
print(f"Saved {OUT}")
for s in wb.sheetnames:
    print(" ·", s)
